from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_
from typing import Optional
from io import BytesIO
from datetime import datetime

from app.core.deps import get_current_user, get_db
from app.models.student import Student
from app.models.broadcast import Broadcast, BroadcastLog
from app.models.journey import JourneyRule, StudentJourney
from app.models.risk_score import RiskScore
from app.models.ticket import Ticket
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/executive")
def executive_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Dashboard executivo — visão geral para diretoria"""

    # === ALUNOS ===
    total_students = db.query(Student).count()
    with_phone = db.query(Student).filter(Student.phone.isnot(None), Student.phone != "").count()
    with_moodle = db.query(Student).filter(Student.moodle_user_id.isnot(None)).count()

    # === FINANCEIRO ===
    financial = db.query(
        Student.financial_status,
        func.count(Student.id)
    ).filter(
        Student.financial_status.isnot(None)
    ).group_by(Student.financial_status).all()

    financial_map = {status: count for status, count in financial}
    total_em_dia = financial_map.get("em_dia", 0)
    total_pendente = financial_map.get("pendente", 0)
    total_inadimplente = financial_map.get("inadimplente", 0)
    total_com_financeiro = total_em_dia + total_pendente + total_inadimplente

    overdue_total = db.query(func.sum(Student.overdue_value)).filter(
        Student.overdue_value > 0
    ).scalar() or 0

    # === RISCO ===
    risk_scores = db.query(
        RiskScore.level,
        func.count(RiskScore.id)
    ).group_by(RiskScore.level).all()

    risk_map = {str(level): count for level, count in risk_scores}

    # === POR CURSO ===
    courses = db.query(
        Student.primary_course_name,
        func.count(Student.id).label("total"),
        func.sum(case((Student.financial_status == "inadimplente", 1), else_=0)).label("inadimplentes"),
        func.sum(case((Student.financial_status == "em_dia", 1), else_=0)).label("em_dia"),
        func.sum(case((Student.financial_status == "pendente", 1), else_=0)).label("pendentes"),
        func.avg(Student.overdue_value).label("avg_overdue"),
    ).filter(
        Student.primary_course_name.isnot(None)
    ).group_by(
        Student.primary_course_name
    ).order_by(
        func.count(Student.id).desc()
    ).all()

    courses_data = []
    for c in courses:
        courses_data.append({
            "course": c[0] if c[0] else "Sem curso",
            "total": c[1],
            "inadimplentes": int(c[2] or 0),
            "em_dia": int(c[3] or 0),
            "pendentes": int(c[4] or 0),
            "avg_overdue": round(float(c[5] or 0), 2),
        })

    # === DOCUMENTAÇÃO ===
    docs_complete = db.query(Student).filter(
        Student.documents_count >= Student.documents_total,
        Student.documents_total > 0
    ).count()
    docs_incomplete = db.query(Student).filter(
        Student.documents_count > 0,
        Student.documents_count < Student.documents_total
    ).count()
    docs_none = db.query(Student).filter(
        Student.documents_count == 0
    ).count()

    # === ACESSO MOODLE ===
    moodle_accessed = db.query(Student).filter(
        Student.moodle_first_access.isnot(None)
    ).count()
    moodle_never = db.query(Student).filter(
        Student.moodle_user_id.isnot(None),
        Student.moodle_first_access.is_(None)
    ).count()

    # === DISPAROS ===
    total_broadcasts = db.query(Broadcast).count()
    total_messages_sent = db.query(func.sum(Broadcast.sent_count)).scalar() or 0
    total_messages_failed = db.query(func.sum(Broadcast.failed_count)).scalar() or 0

    # === RÉGUAS ===
    total_journeys = db.query(JourneyRule).count()
    active_journeys = db.query(JourneyRule).filter(JourneyRule.is_active == True).count()
    students_in_journey = db.query(StudentJourney).filter(StudentJourney.status == "active").count()

    # === TICKETS ===
    total_tickets = db.query(Ticket).count()

    return {
        "summary": {
            "total_students": total_students,
            "with_phone": with_phone,
            "with_moodle": with_moodle,
            "phone_coverage": round(with_phone / total_students * 100, 1) if total_students else 0,
        },
        "financial": {
            "em_dia": total_em_dia,
            "pendente": total_pendente,
            "inadimplente": total_inadimplente,
            "total": total_com_financeiro,
            "overdue_total": round(float(overdue_total), 2),
            "health_rate": round(total_em_dia / total_com_financeiro * 100, 1) if total_com_financeiro else 0,
        },
        "risk": risk_map,
        "documents": {
            "complete": docs_complete,
            "incomplete": docs_incomplete,
            "none": docs_none,
        },
        "moodle": {
            "accessed": moodle_accessed,
            "never_accessed": moodle_never,
        },
        "broadcasts": {
            "total": total_broadcasts,
            "messages_sent": int(total_messages_sent),
            "messages_failed": int(total_messages_failed),
        },
        "journeys": {
            "total": total_journeys,
            "active": active_journeys,
            "students_active": students_in_journey,
        },
        "tickets": {
            "total": total_tickets,
        },
        "courses": courses_data,
    }
@router.get("/executive/export-excel")
def export_executive_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta dashboard executivo em Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Busca mesmos dados do dashboard
    data = executive_dashboard(db=db, current_user=current_user)

    wb = Workbook()

    # === CORES ===
    header_fill = PatternFill(start_color="2A658F", end_color="2A658F", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, size=14, color="27273D")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="2A658F")
    number_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # === ABA 1: RESUMO ===
    ws = wb.active
    ws.title = "Resumo Executivo"
    ws.sheet_properties.tabColor = "2A658F"

    ws.merge_cells("A1:D1")
    ws["A1"] = "CENAT — Relatório Executivo"
    ws["A1"].font = title_font
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="888888")

    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Visão Geral"
    ws[f"A{row}"].font = subtitle_font
    row += 1

    summary_items = [
        ("Total de Alunos", data["summary"]["total_students"]),
        ("Com Telefone", data["summary"]["with_phone"]),
        ("Com Moodle", data["summary"]["with_moodle"]),
        ("Cobertura Telefone", f"{data['summary']['phone_coverage']}%"),
    ]
    for label, value in summary_items:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = number_font
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Financeiro"
    ws[f"A{row}"].font = subtitle_font
    row += 1

    fin_items = [
        ("Em Dia", data["financial"]["em_dia"], green_fill),
        ("Pendente", data["financial"]["pendente"], yellow_fill),
        ("Inadimplente", data["financial"]["inadimplente"], red_fill),
        ("Valor em Atraso", f"R$ {data['financial']['overdue_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), red_fill),
        ("Saúde Financeira", f"{data['financial']['health_rate']}%", green_fill),
    ]
    for label, value, fill in fin_items:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = number_font
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11)
        ws[f"B{row}"].fill = fill
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Risco"
    ws[f"A{row}"].font = subtitle_font
    row += 1

    for level, label, fill in [("RiskLevel.LOW", "Baixo", green_fill), ("RiskLevel.MEDIUM", "Médio", yellow_fill), ("RiskLevel.HIGH", "Alto", red_fill)]:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = data["risk"].get(level, 0)
        ws[f"B{row}"].fill = fill
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=11)
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    # === ABA 2: CURSOS ===
    ws2 = wb.create_sheet("Desempenho por Curso")
    ws2.sheet_properties.tabColor = "4CAF50"

    headers = ["Curso", "Alunos", "Em Dia", "Pendentes", "Inadimplentes", "% Saúde", "Média Atraso (R$)"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, course in enumerate(data["courses"], 2):
        total_fin = course["em_dia"] + course["pendentes"] + course["inadimplentes"]
        health = round((course["em_dia"] / total_fin * 100), 1) if total_fin else 0

        ws2.cell(row=i, column=1, value=course["course"]).border = thin_border
        ws2.cell(row=i, column=2, value=course["total"]).border = thin_border
        ws2.cell(row=i, column=3, value=course["em_dia"]).border = thin_border
        ws2.cell(row=i, column=4, value=course["pendentes"]).border = thin_border
        ws2.cell(row=i, column=5, value=course["inadimplentes"]).border = thin_border
        ws2.cell(row=i, column=6, value=health).border = thin_border
        ws2.cell(row=i, column=7, value=course["avg_overdue"]).border = thin_border

        # Colorir inadimplentes
        if course["inadimplentes"] > 10:
            ws2.cell(row=i, column=5).fill = red_fill
        
        ws2.cell(row=i, column=2).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=3).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=4).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=5).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=6).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=7).number_format = '#,##0.00'

    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 18

    # Salva
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_executivo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
@router.get("/executive/export-pdf")
def export_executive_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta dashboard executivo em PDF"""
    from fpdf import FPDF

    data = executive_dashboard(db=db, current_user=current_user)

    class PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 18)
            self.set_text_color(39, 39, 61)
            self.cell(0, 12, "CENAT - Relatorio Executivo", ln=True)
            self.set_font("Helvetica", "", 9)
            self.set_text_color(120, 120, 120)
            self.cell(0, 6, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
            self.ln(4)
            self.set_draw_color(42, 101, 143)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(6)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")

        def section_title(self, title):
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(42, 101, 143)
            self.cell(0, 10, title, ln=True)
            self.ln(2)

        def kpi_row(self, label, value, color=None):
            self.set_font("Helvetica", "", 10)
            self.set_text_color(80, 80, 80)
            self.cell(70, 7, label)
            self.set_font("Helvetica", "B", 10)
            if color == "green":
                self.set_text_color(16, 124, 65)
            elif color == "red":
                self.set_text_color(200, 40, 40)
            elif color == "amber":
                self.set_text_color(180, 130, 0)
            else:
                self.set_text_color(39, 39, 61)
            self.cell(0, 7, str(value), ln=True)

    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()

    # === VISAO GERAL ===
    pdf.section_title("Visao Geral")
    pdf.kpi_row("Total de Alunos", f"{data['summary']['total_students']:,}".replace(",", "."))
    pdf.kpi_row("Com Telefone", f"{data['summary']['with_phone']:,} ({data['summary']['phone_coverage']}%)".replace(",", "."))
    pdf.kpi_row("Com Moodle", f"{data['summary']['with_moodle']:,}".replace(",", "."))
    pdf.kpi_row("Acessaram Moodle", f"{data['moodle']['accessed']:,}".replace(",", "."))
    pdf.kpi_row("Nunca Acessaram", str(data["moodle"]["never_accessed"]), "red")
    pdf.ln(4)

    # === FINANCEIRO ===
    pdf.section_title("Situacao Financeira")
    pdf.kpi_row("Em Dia", str(data["financial"]["em_dia"]), "green")
    pdf.kpi_row("Pendente", str(data["financial"]["pendente"]), "amber")
    pdf.kpi_row("Inadimplente", str(data["financial"]["inadimplente"]), "red")
    overdue_fmt = f"R$ {data['financial']['overdue_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    pdf.kpi_row("Valor em Atraso", overdue_fmt, "red")
    pdf.kpi_row("Saude Financeira", f"{data['financial']['health_rate']}%", "green")
    pdf.ln(4)

    # === RISCO ===
    pdf.section_title("Distribuicao de Risco")
    pdf.kpi_row("Baixo", str(data["risk"].get("RiskLevel.LOW", 0)), "green")
    pdf.kpi_row("Medio", str(data["risk"].get("RiskLevel.MEDIUM", 0)), "amber")
    pdf.kpi_row("Alto", str(data["risk"].get("RiskLevel.HIGH", 0)), "red")
    pdf.ln(4)

    # === DOCUMENTACAO ===
    pdf.section_title("Documentacao")
    pdf.kpi_row("Completa", str(data["documents"]["complete"]), "green")
    pdf.kpi_row("Incompleta", str(data["documents"]["incomplete"]), "amber")
    pdf.kpi_row("Sem Documentos", str(data["documents"]["none"]), "red")
    pdf.ln(4)

    # === OPERACOES ===
    pdf.section_title("Operacoes")
    pdf.kpi_row("Disparos Realizados", str(data["broadcasts"]["total"]))
    pdf.kpi_row("Mensagens Enviadas", str(data["broadcasts"]["messages_sent"]))
    pdf.kpi_row("Reguas Ativas", str(data["journeys"]["active"]))
    pdf.kpi_row("Alunos em Jornada", str(data["journeys"]["students_active"]))

    # === PAGINA 2: TABELA DE CURSOS ===
    pdf.add_page("L")
    pdf.section_title(f"Desempenho por Curso ({len(data['courses'])} cursos)")

    # Header tabela
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(42, 101, 143)
    pdf.set_text_color(255, 255, 255)
    col_widths = [95, 20, 20, 25, 30, 25, 30, 30]
    headers = ["Curso", "Alunos", "Em Dia", "Pendentes", "Inadimpl.", "% Saude", "Media Atraso"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, fill=True, align="C")
    pdf.ln()

    # Rows
    pdf.set_font("Helvetica", "", 7)
    for j, course in enumerate(data["courses"]):
        if pdf.get_y() > 180:
            pdf.add_page("L")
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(42, 101, 143)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)

        total_fin = course["em_dia"] + course["pendentes"] + course["inadimplentes"]
        health = round((course["em_dia"] / total_fin * 100), 1) if total_fin else 0
        avg_fmt = f"R$ {course['avg_overdue']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        if j % 2 == 0:
            pdf.set_fill_color(245, 247, 250)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.set_text_color(50, 50, 50)
        name = course["course"][:55] + "..." if len(course["course"]) > 55 else course["course"]
        pdf.cell(col_widths[0], 7, name, border=1, fill=True)
        pdf.cell(col_widths[1], 7, str(course["total"]), border=1, fill=True, align="C")

        pdf.set_text_color(16, 124, 65)
        pdf.cell(col_widths[2], 7, str(course["em_dia"]), border=1, fill=True, align="C")

        pdf.set_text_color(180, 130, 0)
        pdf.cell(col_widths[3], 7, str(course["pendentes"]), border=1, fill=True, align="C")

        pdf.set_text_color(200, 40, 40)
        pdf.cell(col_widths[4], 7, str(course["inadimplentes"]), border=1, fill=True, align="C")

        pdf.set_text_color(50, 50, 50)
        pdf.cell(col_widths[5], 7, f"{health}%", border=1, fill=True, align="C")
        pdf.cell(col_widths[6], 7, avg_fmt, border=1, fill=True, align="R")
        pdf.ln()

    buffer = BytesIO()
    pdf.output(buffer)
    buffer.seek(0)

    filename = f"relatorio_executivo_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
@router.get("/inadimplencia")
def inadimplencia_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Relatório de inadimplência por curso"""

    # Totais gerais
    total_students = db.query(Student).filter(Student.financial_status.isnot(None)).count()
    total_inadimplente = db.query(Student).filter(Student.financial_status == "inadimplente").count()
    total_pendente = db.query(Student).filter(Student.financial_status == "pendente").count()
    total_overdue = db.query(func.sum(Student.overdue_value)).filter(Student.overdue_value > 0).scalar() or 0

    # Por curso
    courses = db.query(
        Student.primary_course_name,
        func.count(Student.id).label("total"),
        func.sum(case((Student.financial_status == "em_dia", 1), else_=0)).label("em_dia"),
        func.sum(case((Student.financial_status == "pendente", 1), else_=0)).label("pendentes"),
        func.sum(case((Student.financial_status == "inadimplente", 1), else_=0)).label("inadimplentes"),
        func.sum(case((Student.overdue_value > 0, Student.overdue_value), else_=0)).label("total_overdue"),
        func.max(Student.overdue_value).label("max_overdue"),
    ).filter(
        Student.primary_course_name.isnot(None),
        Student.financial_status.isnot(None),
    ).group_by(
        Student.primary_course_name
    ).order_by(
        func.sum(case((Student.financial_status == "inadimplente", 1), else_=0)).desc()
    ).all()

    courses_data = []
    for c in courses:
        total_fin = int(c[2] or 0) + int(c[3] or 0) + int(c[4] or 0)
        inadimpl_rate = round(int(c[4] or 0) / total_fin * 100, 1) if total_fin else 0
        courses_data.append({
            "course": c[0],
            "total": c[1],
            "em_dia": int(c[2] or 0),
            "pendentes": int(c[3] or 0),
            "inadimplentes": int(c[4] or 0),
            "total_overdue": round(float(c[5] or 0), 2),
            "max_overdue": round(float(c[6] or 0), 2),
            "inadimplencia_rate": inadimpl_rate,
        })

    # Top 10 alunos com maior dívida
    top_debtors = db.query(
        Student.name,
        Student.email,
        Student.phone,
        Student.primary_course_name,
        Student.overdue_value,
        Student.financial_status,
    ).filter(
        Student.overdue_value > 0
    ).order_by(
        Student.overdue_value.desc()
    ).limit(10).all()

    debtors_data = [
        {
            "name": d[0],
            "email": d[1],
            "phone": d[2],
            "course": d[3],
            "overdue_value": round(float(d[4]), 2),
            "status": d[5],
        }
        for d in top_debtors
    ]

    return {
        "summary": {
            "total_students": total_students,
            "total_inadimplente": total_inadimplente,
            "total_pendente": total_pendente,
            "total_overdue": round(float(total_overdue), 2),
            "inadimplencia_rate": round(total_inadimplente / total_students * 100, 1) if total_students else 0,
        },
        "courses": courses_data,
        "top_debtors": debtors_data,
    }
@router.get("/courses")
def courses_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Relatório de desempenho por curso"""
    from app.models.moodle_signal import MoodleSignal

    # Dados gerais por curso (students)
    courses_base = db.query(
        Student.primary_course_name,
        Student.primary_course_id,
        func.count(Student.id).label("total"),
        func.sum(case((Student.financial_status == "em_dia", 1), else_=0)).label("em_dia"),
        func.sum(case((Student.financial_status == "pendente", 1), else_=0)).label("pendentes"),
        func.sum(case((Student.financial_status == "inadimplente", 1), else_=0)).label("inadimplentes"),
        func.sum(case((Student.moodle_first_access.isnot(None), 1), else_=0)).label("acessaram"),
        func.sum(case((and_(Student.moodle_user_id.isnot(None), Student.moodle_first_access.is_(None)), 1), else_=0)).label("nunca_acessaram"),
        func.sum(case((and_(Student.documents_count >= Student.documents_total, Student.documents_total > 0), 1), else_=0)).label("docs_ok"),
    ).filter(
        Student.primary_course_name.isnot(None)
    ).group_by(
        Student.primary_course_name, Student.primary_course_id
    ).order_by(
        func.count(Student.id).desc()
    ).all()

    # Dados Moodle por curso (signals)
    from sqlalchemy import cast, Numeric
    moodle_stats = db.query(
        MoodleSignal.course_id,
        func.round(cast(func.avg(MoodleSignal.progress_percent), Numeric), 1).label("avg_progress"),
        func.round(cast(func.avg(MoodleSignal.days_since_access), Numeric), 0).label("avg_days_since"),
        func.round(cast(func.avg(MoodleSignal.course_grade), Numeric), 1).label("avg_grade"),
        func.count(func.distinct(MoodleSignal.student_id)).label("students_with_data"),
    ).group_by(MoodleSignal.course_id).all()

    moodle_map = {}
    for m in moodle_stats:
        moodle_map[m[0]] = {
            "avg_progress": float(m[1] or 0),
            "avg_days_since": int(m[2] or 0),
            "avg_grade": float(m[3] or 0),
            "students_with_data": m[4],
        }

    # Risk por curso
    from app.models.risk_score import RiskScore
    risk_stats = db.query(
        Student.primary_course_id,
        RiskScore.level,
        func.count(RiskScore.id),
    ).join(
        RiskScore, RiskScore.student_id == Student.id
    ).filter(
        Student.primary_course_id.isnot(None)
    ).group_by(
        Student.primary_course_id, RiskScore.level
    ).all()

    risk_map = {}
    for r in risk_stats:
        cid = r[0]
        if cid not in risk_map:
            risk_map[cid] = {"low": 0, "medium": 0, "high": 0}
        level_str = str(r[1]).replace("RiskLevel.", "").lower()
        risk_map[cid][level_str] = r[2]

    # Monta resultado
    courses_data = []
    total_students = 0
    total_progress = 0
    courses_with_progress = 0

    for c in courses_base:
        course_id = c[1]
        moodle = moodle_map.get(course_id, {})
        risk = risk_map.get(course_id, {"low": 0, "medium": 0, "high": 0})
        total_fin = int(c[3] or 0) + int(c[4] or 0) + int(c[5] or 0)
        health_rate = round(int(c[3] or 0) / total_fin * 100, 1) if total_fin else 0

        avg_progress = moodle.get("avg_progress", 0)
        if avg_progress > 0:
            total_progress += avg_progress
            courses_with_progress += 1

        total_students += c[2]

        courses_data.append({
            "course": c[0],
            "course_id": course_id,
            "total": c[2],
            "financial": {
                "em_dia": int(c[3] or 0),
                "pendentes": int(c[4] or 0),
                "inadimplentes": int(c[5] or 0),
                "health_rate": health_rate,
            },
            "moodle": {
                "acessaram": int(c[6] or 0),
                "nunca_acessaram": int(c[7] or 0),
                "avg_progress": avg_progress,
                "avg_days_since": moodle.get("avg_days_since", 0),
                "avg_grade": moodle.get("avg_grade", 0),
            },
            "docs_ok": int(c[8] or 0),
            "risk": risk,
        })

    return {
        "summary": {
            "total_courses": len(courses_data),
            "total_students": total_students,
            "avg_progress": round(total_progress / courses_with_progress, 1) if courses_with_progress else 0,
        },
        "courses": courses_data,
    }